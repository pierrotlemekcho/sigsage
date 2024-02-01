from pathlib import Path
import pickle 
from openpyxl import Workbook

HERE = Path(__file__).parent
DATA_FOLDER = HERE / "data"

wb = Workbook()
ws = wb.active


with open(DATA_FOLDER / "tableau.pkl", "rb") as tf:
    dic_compte = pickle.load(tf)

lcompte = list(dic_compte.keys())
print(lcompte)
for i in range(len(lcompte)):
    print("indice"+str(i)+", --->"+lcompte[i])
lgv1=['Ventes prestations',
      lcompte[52],
      lcompte[53],
      lcompte[54],
      lcompte[55],
      lcompte[56],
      lcompte[57],
      lcompte[58],
      lcompte[59],
      lcompte[60],
      lcompte[61],
      lcompte[62],
      lcompte[63]]
lgv2=['Tansfert de charges',
      lcompte[64],
      lcompte[65],
      lcompte[66]]
lg1=['Achat Matière',
     lcompte[0],
     lcompte[1],
     lcompte[2],
     lcompte[3]]
lg2=['Sous traitance' ,
     lcompte[4]]
lg3=['Couts diects chantiers',
     lcompte[9],
     lcompte[5],
     lcompte[16],
     lcompte[12],
     lcompte[6],
     lcompte[25],
     lcompte[32],
     lcompte[7]]
lg4=['Salaire personnel de chantier',
     lcompte[36]]
lg5=['Charges salaire de chantier',
     lcompte[33],
     lcompte[38],
     lcompte[39],
     lcompte[41],
     lcompte[42],
     lcompte[44],
     lcompte[45],
     lcompte[34]]
lg6=['Salaires direction',
     lcompte[37]]
lg7=['Charges salaires de direcion',
    lcompte[40],
    lcompte[43],
    lcompte[46],
    lcompte[47]]
lg8=['Frais généraux',
     lcompte[27],
     lcompte[29],
     lcompte[28],
     lcompte[11],
     lcompte[14]]





lgroupes=[lgv1,lgv2,lg1,lg2,lg3,lg4,lg5,lg7,lg8]
print(lgroupes)

#p#rint('########################')
#print(lg1)
#dg1 = {
#    "titre": lg1[0],
#    lg1[1] : dic_compte.get(lg1[1]),
#    lg1[2] : dic_compte.get(lg1[2]),
#    lg1[3] : dic_compte.get(lg1[3]),
#}
# boucle sur la liste de groupes lgroupes (ii indice ligne , jj indice colonne
# d'excel)

dict_compte_selec={}
dic_compte_solde={}
ii=1
for lgroup in lgroupes:
    dgroup = {}
    dgroup = {"titre" : lgroup[0] }
    for i in range(1,len(lgroup)):
        dgroup[lgroup[i]] = dic_compte.get(lgroup[i])
        dict_compte_selec[lgroup[i]] = dic_compte.get(lgroup[i])

        #print('########################')
        #print(dg1)
        #print('########################')
        #print(dic_compte.get(lg1[1]).items())
        #print(dic_compte.get(lg1[1]).keys())
        #print(dic_compte.get(lg1[1]).values())
        #print('########################')
        #print('########################')
        #print(dic_compte.get(lg1[1]).get("mois4"))


    stotal = {}
    for i in range(1,len(lgroup)):
        if i == 1 :
            for key, value in dic_compte.get(lgroup[i]).items():
                stotal[key] =  value
                print(value)
    #            print('########################')
    #            print(stotal)
    #            print('########################')
        else :
            for key, value in dic_compte.get(lgroup[i]).items():
                stotal[key] = stotal[key] + value
    #            print('########################')
    #            print(stotal)
    #            print('########################')
    print('~~~~~~~~~~~~~~~~~~')
    print(i)
    print('~~~~~~~~~~~~~~~~~~')

    print('#####ws###################')
    print(stotal)
    print('########################')

    dgroup[str("Sous-total : "+lgroup[0])] = stotal

    print(dgroup)

    #ii=1
    jj=1

    #ws.cell(row=ii, column=jj, value = lg1[0])

    for key, value in dgroup.items():
        if isinstance(value, str):
            ws.cell(row=ii, column=jj, value =value)
            print(key,value)
        else:
            ii+=1
            jj=1
            ws.cell(row=ii, column=jj, value = key)

            for key1, value1 in value.items():
                jj+=1
                print(key1 ,value1)
                ws.cell(row=ii, column=jj, value =value1)

            print( value)
            print(type(value))
    ii+=2
    
ii+=2
jj=1
ws.cell(row=ii, column=jj, value ='compte NON affectés')

for key, value in dic_compte.items():
    if key not in dict_compte_selec:
        ii+=1
        jj=1
        lgroupes.append(key) #maj groupes avec comptes non affectées
        ws.cell(row=ii, column=jj, value = key)
        for key1, value1 in value.items():
            jj+=1
            print(key1 ,value1)
            ws.cell(row=ii, column=jj, value =value1)


        





print(dic_compte_solde)


#ws['A1'] = lg1[0]

wb.save(DATA_FOLDER / 'document.xlsx')
with open(DATA_FOLDER / "groupes.pkl", "wb") as tf1:
    pickle.dump(lgroupes , tf1)

